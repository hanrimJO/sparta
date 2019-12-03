from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

work_book = load_workbook('./prac01.xlsx')
work_sheet = work_book['prac']


headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://movie.naver.com/movie/sdb/rank/rmovie.nhn?sel=pnt&date=20190909',headers=headers)

soup = BeautifulSoup(data.text, 'html.parser')

# select를 이용해서, tr들을 불러오기
movies = soup.select('#old_content > table > tbody > tr')

# soup.select('input[type="checkbox"]')

# movies (tr들) 의 반복문을 돌리기

num = 0
for movie in movies:
    # movie 안에 a 가 있으면,

    a_tag = movie.select_one('td.title > div > a')
    rank = movie.select_one('td.point')

    if a_tag is not None:
        # a의 text를 찍어본다.
        num += 1
        # print(a, end=' ')
        print(num, a_tag.text, rank.text)
        work_sheet.cell(row=num+1, column=1, value=num)
        work_sheet.cell(row=num+1, column=2, value=a_tag.text)
        work_sheet.cell(row=num+1, column=3, value=rank.text)





# 수정한 엑셀파일을 저장합니다.
# 참고: 다른이름으로 저장할 수도 있습니다.
work_book.save('./prac01.xlsx')